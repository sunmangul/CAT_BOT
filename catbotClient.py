import datetime
import discord
import openpyxl

from random import *

client = discord.Client()

token = "NzU1NDI5OTc0MjgyMDc2MjUy.X2DK_Q.I_OXm-2sphKyHiLJqudDxdeedbo"

@client.event
async def on_ready():
    print("시작")
    game = discord.Game("다이어트")
    await client.change_presence(status=discord.Status.online, activity=discord.Game("다이어트"))

@client.event
async def on_message(message):
    if message.author.bot:
        return None
    # if message.content.startswith("단단아"):
    #     await message.channel.send("불렀..?")
    elif message.content.startswith("단단아 안녕"):
        await message.channel.send("그래... 뭐 안녕")
    elif message.content.startswith("단단아 에이"):
        await message.channel.send("에이야 단단이")
    # if message.content.startswith("단단아 굴러"):
    #     await message.channel.send("ㄷ..데ㄱ...(무거워서 안굴러간다)")
    elif message.content.startswith("단단아 굴러"):
        await message.channel.send("데굴데굴(하는척,,,)")
    elif message.content.startswith("단단아 팀장"):
        await message.channel.send("먀아ㅏ")
    elif message.content.startswith("단단아 도움말"):
        embed = discord.Embed(title="단단이:", description="단단이의 기본적인 커맨드를 보여줍니다.", color=0xd6e6d, timestamp=datetime.datetime.utcnow())
        embed.add_field(name="굴러:", value="단단이가 혼신의 힘을 다해 구르려고 합니다", inline=False)
        embed.add_field(name="에이:", value="단단이의 봇 친구 에이를 부릅니다..", inline=False)
        embed.add_field(name="팀장:", value="우리팀의 팀장을 부릅니다..", inline=False)
        embed.add_field(name="배워:", value="""단단이에게 단어를 가르칠 수 있습니다
                                              ex)단단아 배워 단어:내용""", inline=False)
        await message.channel.send(embed=embed)
    elif message.content.startswith("단단아 자기소개"):
        embed = discord.Embed(title="단단이:", description="말랑이 후속으로 말랑이 보다 똑똑하고, 더 단단해졌습니다.", color=0xd6e6d,
                              timestamp=datetime.datetime.utcnow())
        embed.add_field(name="이름:", value="단단한 고양이", inline=False)
        embed.add_field(name="나이:", value="1살.", inline=False)
        embed.add_field(name="생일:", value="2020년 9월 15일.", inline=False)
        await message.channel.send(embed=embed)

    elif message.content.startswith("단단아 주사위"):
        i = randint(1, 6)
        await message.channel.send("데구르르... %d" %i)

    elif message.content.startswith("단단아 큐브"):
        i = randint(1, 8)
        file = openpyxl.load_workbook("큐브.xlsx")
        sheet = file.active
        await message.channel.send("오늘은 " + sheet["A" + str(i)].value + " 맞춰봐")

    elif message.content.startswith("단단아 배워"):
        i = 0
        file = openpyxl.load_workbook("리스트.xlsx")
        sheet = file.active
        learn1 = message.content.split(':')
        learn2 = learn1[0].replace("단단아 배워 ", "")
        while True:
            i += 1
            if sheet["A" + str(i)].value == None:
                await message.channel.send("음... 그래")
                sheet["A" + str(i)].value = learn2
                sheet["B" + str(i)].value = learn1[1]
                break
            elif sheet["A" + str(i)].value == learn2:
                await message.channel.send("이미 있는 단어야")
                break
        file.save("리스트.xlsx")

    elif message.content.startswith("단단아 잊어"):
        i = 0
        file = openpyxl.load_workbook("리스트.xlsx")
        sheet = file.active
        learn = message.content.replace('단단아 잊어 ', "")
        while True:
            i += 1
            if sheet["A" + str(i)].value == learn:
                await message.channel.send("귀찮게....")
                sheet["A" + str(i)].value = ""
                sheet["B" + str(i)].value = ""
                break
            if sheet["A" + str(i)].value == None:
                await message.channel.send("그런 단어가 있나..?")
                break
        file.save("리스트.xlsx")

    elif message.content.startswith("단단아 "):
        i = 0
        file = openpyxl.load_workbook("리스트.xlsx")
        sheet = file.active
        memory = message.content.replace("단단아 ", "")
        while True:
            i += 1
            if sheet["A" + str(i)].value == memory:
                await message.channel.send(sheet["B" + str(i)].value)
                break
            if sheet["A" + str(i)].value == None:
                await message.channel.send("그런거 배운적 없는데..?")
                break



    # if message.content.startswith("단단아 기억해"):
    #     file = open("리스트.txt", 'w')
    #     txt = file.active
    #     learn = message.content.split(' ')
    #     for i in range(1, 51):
    #         file.write(txt)
    #     file.close() 메모장 실패



client.run(token)
